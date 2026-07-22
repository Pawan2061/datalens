import io
import re
import smtplib
import ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr, parseaddr

import openpyxl

from app.config import settings

_SKIP_CHART_TYPES = {"kpi", "gauge"}
_EMAIL_IN_TEXT_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
_EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")


class EmailDeliveryError(RuntimeError):
    """Raised when an SMTP message cannot be delivered to all recipients."""


def build_excel_bytes(tables: list[dict], charts: list[dict]) -> bytes | None:
    """Build an xlsx workbook from InsightResult tables and chart data.

    Returns raw bytes, or None if there is nothing exportable.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # discard the default empty sheet

    for i, table in enumerate(tables):
        title = _sanitize_sheet_name(table.get("title") or f"Table {i + 1}")
        ws = wb.create_sheet(title=title)
        cols: list[str] = table.get("columns") or []
        ws.append(cols)
        for row in table.get("data") or []:
            ws.append([row.get(c) for c in cols])

    for i, chart in enumerate(charts):
        if chart.get("chart_type") in _SKIP_CHART_TYPES:
            continue
        rows: list[dict] = chart.get("data") or []
        if not rows:
            continue
        cols = list(rows[0].keys())
        title = _sanitize_sheet_name(chart.get("title") or f"Chart {i + 1}")
        ws = wb.create_sheet(title=title)
        ws.append(cols)
        for row in rows:
            ws.append([row.get(c) for c in cols])

    if not wb.sheetnames:
        return None

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_excel_email(
    *,
    to: str,
    subject: str,
    body_html: str,
    excel_bytes: bytes,
    filename: str,
) -> None:
    """Send an email with an xlsx attachment using the configured SMTP settings."""
    recipients = [(to or "").strip()]
    _validate_recipients(recipients)
    header_from, envelope_from = _smtp_from_values()
    msg = MIMEMultipart()
    msg["From"] = header_from
    msg["To"] = recipients[0]
    msg["Subject"] = subject
    msg.attach(MIMEText(body_html, "html"))

    part = MIMEBase(
        "application",
        "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    part.set_payload(excel_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    _send_message(envelope_from, recipients, msg)


def send_alert_email(*, to: list[str], subject: str, body_html: str) -> None:
    """Send a plain HTML alert email (no attachment) to one or more recipients.

    Uses the same SMTP settings/flow as send_excel_email(). Blocking — call
    from a thread (asyncio.to_thread) when on the event loop.
    """
    recipients = [(recipient or "").strip() for recipient in to]
    _validate_recipients(recipients)
    header_from, envelope_from = _smtp_from_values()
    msg = MIMEMultipart()
    msg["From"] = header_from
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = subject
    msg.attach(MIMEText(body_html, "html"))

    _send_message(envelope_from, recipients, msg)


def _validate_recipients(recipients: list[str]) -> None:
    if not recipients or any(not _EMAIL_RE.fullmatch((recipient or "").strip()) for recipient in recipients):
        raise EmailDeliveryError("No valid email recipients were provided.")


def _send_message(envelope_from: str, recipients: list[str], message: MIMEMultipart) -> None:
    host = (settings.email_smtp_host or "").strip()
    username = (settings.email_smtp_user or "").strip()
    password = settings.email_smtp_pass or ""
    if not host or not username or not password:
        raise EmailDeliveryError("SMTP is not configured. Set EMAIL_SMTP_HOST, EMAIL_SMTP_USER, and EMAIL_SMTP_PASS.")
    if not envelope_from:
        raise EmailDeliveryError("SMTP sender is not configured.")

    timeout = max(1, int(settings.email_smtp_timeout_seconds or 30))
    context = ssl.create_default_context()
    smtp_cls = smtplib.SMTP_SSL if settings.email_smtp_use_ssl else smtplib.SMTP
    try:
        with smtp_cls(settings.email_smtp_host, settings.email_smtp_port, timeout=timeout) as server:
            server.ehlo()
            if settings.email_smtp_use_starttls and not settings.email_smtp_use_ssl:
                server.starttls(context=context)
                server.ehlo()
            server.login(username, password)
            refused = server.sendmail(envelope_from, recipients, message.as_string())
    except (OSError, smtplib.SMTPException) as exc:
        raise EmailDeliveryError(f"SMTP delivery failed: {exc}") from exc

    if refused:
        refused_addresses = ", ".join(sorted(refused))
        raise EmailDeliveryError(f"SMTP rejected recipient(s): {refused_addresses}")


def _smtp_from_values() -> tuple[str, str]:
    """Return RFC-safe header From plus plain envelope sender for SMTP MAIL FROM.

    Gmail requires the envelope sender to be the authenticated mailbox (or an
    explicitly allowed alias). Keep EMAIL_SMTP_USER as the authoritative SMTP
    sender and treat EMAIL_SMTP_FROM as an optional display/header value only.
    """
    user_email = (settings.email_smtp_user or "").strip()
    raw_from = (settings.email_smtp_from or settings.email_smtp_user or "").strip()
    display_name, parsed_email = parseaddr(raw_from)
    extracted = _EMAIL_IN_TEXT_RE.search(raw_from)
    email_addr = parsed_email if _EMAIL_IN_TEXT_RE.fullmatch(parsed_email or "") else ""
    email_addr = email_addr or (extracted.group(0) if extracted else raw_from)
    envelope_from = user_email or email_addr
    if display_name and envelope_from:
        return formataddr((display_name, envelope_from)), envelope_from
    return envelope_from, envelope_from


def _sanitize_sheet_name(name: str) -> str:
    """Remove chars Excel forbids in sheet names and cap at 31 characters."""
    cleaned = name.translate(str.maketrans("", "", r'\/?*[]:' )).strip()
    return (cleaned or "Sheet")[:31]
